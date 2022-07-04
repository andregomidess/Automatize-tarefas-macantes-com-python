import openpyxl


planilha = openpyxl.load_workbook('produceSales.xlsx')
sheet = planilha.active
atualiza_precos = {'Garlic': 3.07, 'Celery': 1.19, 'Lemon': 1.27}

for linhas in range(2, sheet.max_row):
    nome_produto = sheet.cell(row = linhas, column=1).value
    if nome_produto in atualiza_precos:
        sheet.cell(row=linhas, column=1).value = atualiza_precos[nome_produto]
planilha.save('PlanilhaPrecosAtualizados.xlsx')        

