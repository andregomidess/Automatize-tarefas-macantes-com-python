import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string


wb = openpyxl.load_workbook('example.xlsx')
sheet = wb['Sheet1']
sheet['A1']
print(sheet['A1'].value)
c = sheet['B1']
print(c.value)
print(f'Linha {c.row}, coluna {c.column} é {c.value}')
print(f'Célula {c.coordinate} é {c.value}')
print(sheet['C1'].value)

sheet.cell(row=1, column=2) # metodo cell pega as linha e colunas
sheet.cell(row=1, column=2).value # metodo cell().value pega o valor que está nessa linha e coluna
for i in range(1, 8, 2):
    print(i, sheet.cell(row=i, column=2).value)
    
max_colunas = sheet.max_column   
print(f'número de colunas no documento: {max_colunas}')
max_linhas = sheet.max_row
print(f'Número máximo de linhas no documento: {max_linhas}')

print(get_column_letter(1)) # converte a o n° da coluna em letra (padrão excel)
print(get_column_letter(2))
print(get_column_letter(27))
print(get_column_letter(8997))

print(get_column_letter(sheet.max_column)) #pegando o máximo de colunas do sheet1 e convertendo em letra
print(column_index_from_string('A'))    # converte a letra em n°
print(column_index_from_string('AA'))

#O loop for externo percorre todas as linhas do slice u. Então, para cada linha, o loop for aninhado percorre todas as células dessa linha v.
print(tuple(sheet['A1':'C3']))
for linhas_celula in sheet['A1':'C3']:  #u
    for cellObj in linhas_celula:       #v
        print(cellObj.coordinate, cellObj.value)
    print('--- FIM DA LINHA ---')
    
sheet = wb.active
print(list(sheet.columns)[1])
for obj_celula in list(sheet.columns)[1]:   # laço for para percorrer so as colunas, através do metodo columns
    print(obj_celula.value)  

print(list(sheet.rows))    
for obj_celula in list(sheet.rows)[1]:   # laço for para percorrer so as colunas, através do metodo columns
    print(obj_celula)